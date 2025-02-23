from flask import Flask, render_template, request, jsonify, send_file
from datetime import datetime
import openai
from dotenv import load_dotenv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io

app = Flask(__name__)

# Load environment variables
load_dotenv()

# Configure OpenAI
openai.api_key = os.getenv('OPENAI_API_KEY')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/rewrite_text', methods=['POST'])
def rewrite_text():
    content = request.json.get('content')
    if not content:
        return jsonify({'error': 'No content provided'}), 400

    try:
        response = openai.ChatCompletion.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": "Rewrite the following text professionally and organize it in bullet points:"},
                {"role": "user", "content": content}
            ]
        )
        rewritten_text = response.choices[0].message.content
        return jsonify({'rewritten_text': rewritten_text})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/export_to_excel', methods=['POST'])
def export_to_excel():
    data = request.json

    x = datetime.now()
    filename = f"Daily-MIS-{x.strftime('%x').replace('/', '-')}.xlsx"

    wb = Workbook()
    ws = wb.active
    thick_border = Border(
        left=Side(border_style="medium", color="000000"),
        right=Side(border_style="medium", color="000000"),
        top=Side(border_style="medium", color="000000"),
        bottom=Side(border_style="medium", color="000000")
    )

    headers = ['Date / Day', 'Total Number of Staff', 'No. of Presents', 'No. of Absents',
               'Total Students', 'No. of Presents', 'No. of Absents', 'Update', 'Issues/Concerns']
    header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left')
        cell.fill = header_fill
        cell.border = thick_border

    next_row = 2
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thick_border

    try:
        date_obj = datetime.strptime(data['date'], '%Y-%m-%d')
        day_name = date_obj.strftime('%A')
    except ValueError as ve:
        return jsonify({'error': f"Invalid date format: {str(ve)}"}), 400

    day_date = [f"{date_obj.strftime('%m/%d/%Y')}", f"{day_name}"]

    total_number_of_staff = []
    no_of_presents = []
    staff_present = 0
    total_staff = 0

    for label, total, present in data['staff']:
        total = int(total or 0)
        present = int(present or 0)
        if total > 0 or present > 0:
            total_number_of_staff.append(f"{label}: {total}")
            no_of_presents.append(f"{label}: {present}")
            staff_present += present
            total_staff += total

    total_Staff_Absent = total_staff - staff_present

    total_student_attendance = []
    total_student_presents = []

    for label, total, present in data['students']:
        total = int(total or 0)
        present = int(present or 0)
        if total > 0 or present > 0:
            total_student_attendance.append(f"{label}: {total}")
            total_student_presents.append(f"{label}: {present}")

    total_present = sum(int(present or 0) for _, _, present in data['students'])
    total_students = sum(int(total or 0) for _, total, _ in data['students'])
    total_absent = total_students - total_present

    total_student_presents.append(f"Total: {total_present}")

    data = [
        "\n".join(day_date),
        "\n".join(total_number_of_staff),
        "\n".join(no_of_presents),
        total_Staff_Absent,
        "\n".join(total_student_attendance),
        "\n".join(total_student_presents),
        total_absent,
        data['remarks'],
        data['issues']
    ]

    for col, value in enumerate(data, 1):
        cell = ws.cell(row=next_row, column=col)
        cell.value = value
        cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell.border = thick_border

    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(debug=True)
