import tkinter as tk
from tkinter import ttk, messagebox,filedialog
from tkcalendar import DateEntry
from datetime import datetime
import openai
from dotenv import load_dotenv
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment,PatternFill, Border, Side
import os.path


# Load environment variables
load_dotenv()

# Configure OpenAI
openai.api_key = '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'

class MISCreator:
    def __init__(self, root):
        self.root = root
        self.root.title("MIS Creator")
        self.root.geometry("1000x900")
        
        self.save_folder = None

        # Create main frame with scrollbar
        main_frame = ttk.Frame(root)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(main_frame)
        scrollbar = ttk.Scrollbar(main_frame, orient="vertical", command=canvas.yview)
        self.scrollable_frame = ttk.Frame(canvas)

        self.scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        main_frame.pack(fill=tk.BOTH, expand=True)
        canvas.pack(side="left", fill=tk.BOTH, expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Section to select folder
        folder_frame = ttk.LabelFrame(self.scrollable_frame, text="Export Options", padding=10)
        folder_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(folder_frame, text="Save Folder:").grid(row=0, column=0, padx=5, pady=5)
        self.folder_label = ttk.Label(folder_frame, text="No folder selected", relief="sunken", anchor="w")
        self.folder_label.grid(row=0, column=1, padx=5, pady=5, sticky="ew")
        folder_frame.columnconfigure(1, weight=1)

        ttk.Button(folder_frame, text="Select Folder", command=self.select_folder).grid(row=0, column=2, padx=5, pady=5)

    def select_folder(self):
        """Open a folder selection dialog and save the path."""
        folder = filedialog.askdirectory(title="Select Folder to Save Excel File")
        if folder:
            self.save_folder = folder
            self.folder_label.config(text=folder)
        else:
            self.folder_label.config(text="No folder selected")

        # Date Section
        date_frame = ttk.LabelFrame(self.scrollable_frame, text="Date Information", padding=10)
        date_frame.pack(fill="x", padx=10, pady=5)

        self.date_entry = DateEntry(date_frame, width=12, background='darkblue',
                                  foreground='white', borderwidth=2)
        self.date_entry.grid(row=0, column=1, padx=5)
        
        # Staff Section
        staff_frame = ttk.LabelFrame(self.scrollable_frame, text="Staff Attendance", padding=10)
        staff_frame.pack(fill="x", padx=10, pady=5)

        # Teachers
        ttk.Label(staff_frame, text="Teachers Present:").grid(row=0, column=0, padx=5, pady=5)
        self.teachers_present = ttk.Entry(staff_frame, width=10)
        self.teachers_present.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total Teachers:").grid(row=0, column=2, padx=5, pady=5)
        self.total_teachers = ttk.Entry(staff_frame, width=10)
        self.total_teachers.grid(row=0, column=3, padx=5, pady=5)

        # Admins
        ttk.Label(staff_frame, text="Admins Present:").grid(row=1, column=0, padx=5, pady=5)
        self.admins_present = ttk.Entry(staff_frame, width=10)
        self.admins_present.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total Admins:").grid(row=1, column=2, padx=5, pady=5)
        self.total_admins = ttk.Entry(staff_frame, width=10)
        self.total_admins.grid(row=1, column=3, padx=5, pady=5)

        # House Keeping
        ttk.Label(staff_frame, text="House Keeping Present:").grid(row=2, column=0, padx=5, pady=5)
        self.hks_present = ttk.Entry(staff_frame, width=10)
        self.hks_present.grid(row=2, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total House Keeping:").grid(row=2, column=2, padx=5, pady=5)
        self.total_hks = ttk.Entry(staff_frame, width=10)
        self.total_hks.grid(row=2, column=3, padx=5, pady=5)

        # Teacher Helpers
        ttk.Label(staff_frame, text="Teacher Helpers Present:").grid(row=3, column=0, padx=5, pady=5)
        self.Teacher_helpers = ttk.Entry(staff_frame, width=10)
        self.Teacher_helpers.grid(row=3, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total Teacher Helpers:").grid(row=3, column=2, padx=5, pady=5)
        self.total_Teacher_helpers = ttk.Entry(staff_frame, width=10)
        self.total_Teacher_helpers.grid(row=3, column=3, padx=5, pady=5)
        
        # Security
        ttk.Label(staff_frame, text="Security  Present:").grid(row=4, column=0, padx=5, pady=5)
        self.Security_present = ttk.Entry(staff_frame, width=10)
        self.Security_present.grid(row=4, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total Security:").grid(row=4, column=2, padx=5, pady=5)
        self.total_security = ttk.Entry(staff_frame, width=10)
        self.total_security.grid(row=4, column=3, padx=5, pady=5)
        
        # School Attendant Trainee
        ttk.Label(staff_frame, text="School Attendant Trainee Present:").grid(row=5, column=0, padx=5, pady=5)
        self.attendant_trainee_present = ttk.Entry(staff_frame, width=10)
        self.attendant_trainee_present.grid(row=5, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total School Attendant Trainee:").grid(row=5, column=2, padx=5, pady=5)
        self.total_attendant_trainee = ttk.Entry(staff_frame, width=10)
        self.total_attendant_trainee.grid(row=5, column=3, padx=5, pady=5)

        # School Nurse
        ttk.Label(staff_frame, text="School Nurse Present:").grid(row=6, column=0, padx=5, pady=5)
        self.nurse_present = ttk.Entry(staff_frame, width=10)
        self.nurse_present.grid(row=6, column=1, padx=5, pady=5)
        
        ttk.Label(staff_frame, text="Total School Nurse:").grid(row=6, column=2, padx=5, pady=5)
        self.total_nurse = ttk.Entry(staff_frame, width=10)
        self.total_nurse.grid(row=6, column=3, padx=5, pady=5)
        

        # Students Section
        students_frame = ttk.LabelFrame(self.scrollable_frame, text="Students Attendance", padding=10)
        students_frame.pack(fill="x", padx=10, pady=5)

        # Pre-Primary
        ttk.Label(students_frame, text="Pre-Primary Present:").grid(row=0, column=0, padx=5, pady=5)
        self.preprimary_present = ttk.Entry(students_frame, width=10)
        self.preprimary_present.grid(row=0, column=1, padx=5, pady=5)
        
        ttk.Label(students_frame, text="Total Pre-Primary:").grid(row=0, column=2, padx=5, pady=5)
        self.total_preprimary = ttk.Entry(students_frame, width=10)
        self.total_preprimary.grid(row=0, column=3, padx=5, pady=5)

        # Primary
        ttk.Label(students_frame, text="Primary Present:").grid(row=1, column=0, padx=5, pady=5)
        self.primary_present = ttk.Entry(students_frame, width=10)
        self.primary_present.grid(row=1, column=1, padx=5, pady=5)
        
        ttk.Label(students_frame, text="Total Primary:").grid(row=1, column=2, padx=5, pady=5)
        self.total_primary = ttk.Entry(students_frame, width=10)
        self.total_primary.grid(row=1, column=3, padx=5, pady=5)

        # Total Students Summary
        summary_frame = ttk.Frame(students_frame)
        summary_frame.grid(row=2, column=0, columnspan=4, pady=10)
        
        self.total_present_label = ttk.Label(summary_frame, text="Total Present: 0")
        self.total_present_label.pack(side=tk.LEFT, padx=10)
        
        self.total_absent_label = ttk.Label(summary_frame, text="Total Absent: 0")
        self.total_absent_label.pack(side=tk.LEFT, padx=10)

        # Bind the calculation to entry widgets
        entries = [self.preprimary_present, self.total_preprimary,
                  self.primary_present, self.total_primary]
        for entry in entries:
            entry.bind('<KeyRelease>', self.calculate_totals)

        # Remarks Section
        remarks_frame = ttk.LabelFrame(self.scrollable_frame, text="Update", padding=10)
        remarks_frame.pack(fill="x", padx=10, pady=5)

        self.remarks_text = tk.Text(remarks_frame, height=4, width=50)
        self.remarks_text.pack(fill="x", pady=5)
        
        ttk.Button(remarks_frame, text="Rewrite Updates", 
                  command=lambda: self.rewrite_text(self.remarks_text)).pack(pady=5)

        # Issues Section
        issues_frame = ttk.LabelFrame(self.scrollable_frame, text="Issues / Concerns", padding=10)
        issues_frame.pack(fill="x", padx=10, pady=5)

        self.issues_text = tk.Text(issues_frame, height=4, width=50)
        self.issues_text.pack(fill="x", pady=5)
        
        ttk.Button(issues_frame, text="Rewrite Issues / Concerns", 
                  command=lambda: self.rewrite_text(self.issues_text)).pack(pady=5)

        # Export Button
        ttk.Button(self.scrollable_frame, text="Export to Excel", 
                  command=self.export_to_excel).pack(pady=10)

    def calculate_totals(self, event=None):
        try:
            preprimary_present = int(self.preprimary_present.get() or 0)
            total_preprimary = int(self.total_preprimary.get() or 0)
            primary_present = int(self.primary_present.get() or 0)
            total_primary = int(self.total_primary.get() or 0)


            total_present = preprimary_present + primary_present
            total_students = total_preprimary + total_primary
            total_absent = total_students - total_present

            self.total_present_label.config(text=f"Total Present: {total_present}")
            self.total_absent_label.config(text=f"Total Absent: {total_absent}")
        except ValueError:
            pass

    def rewrite_text(self, text_widget):
        content = text_widget.get("1.0", tk.END).strip()
        if not content:
            messagebox.showwarning("Warning", "Please enter some text to rewrite")
            return

        try:
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "Rewrite the following text professionally and organize it in bullet points:"},
                    {"role": "user", "content": content}
                ]
            )
            
            rewritten_text = response.choices[0].message.content
            text_widget.delete("1.0", tk.END)
            text_widget.insert("1.0", rewritten_text)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to rewrite text: {str(e)}")

    ...
    def export_to_excel(self):
        if not self.save_folder:
            messagebox.showwarning("Warning", "Please select a folder to save the Excel file!")
            return
        
        x = datetime.now()
        # Replace invalid characters in the filename
        filename = f"Daily-MIS-{x.strftime('%x').replace('/', '-')}.xlsx"
        full_path = f"{self.save_folder}/{filename}"
        
        wb = Workbook()
        ws = wb.active
        thick_border = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000")
        )
    
        # Add headers
        headers = ['Date / Day', 'Total Number of Staff', 'No. of Presents', 'No. of Absents',
                   'Total Students', 'No. of Presents', 'No. of Absents', 'Update', 'Issues/Concerns']
        header_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")  # Yellow fill color

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='left')
            cell.fill = header_fill
            cell.border = thick_border
        next_row = 2
        
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thick_border
    
        # Get day name from date
        try:
            date_obj = datetime.strptime(self.date_entry.get(), '%m/%d/%y')
            day_name = date_obj.strftime('%A')
        except ValueError as ve:
            messagebox.showerror("Error", f"Invalid date format: {str(ve)}")
            return
        
        day_date = [
            f"{date_obj.strftime('%m/%d/%Y')}",
            f"{day_name}"
        ]
    
        # Format staff attendance
        staff_labels = [
            ("Teachers", self.total_teachers, self.teachers_present),
            ("Admins", self.total_admins, self.admins_present),
            ("House Keeping", self.total_hks, self.hks_present),
            ("Teacher Helpers", self.total_Teacher_helpers, self.Teacher_helpers),
            ("Security", self.total_security, self.Security_present),
            ("School Attendant Trainee", self.total_attendant_trainee, self.attendant_trainee_present),
            ("School Nurse", self.total_nurse, self.nurse_present)
        ]
        
        total_number_of_staff = []
        no_of_presents = []
        staff_present = 0
        total_staff = 0
        
        for label, total_entry, present_entry in staff_labels:
            try:
                total = int(total_entry.get() or 0)
                present = int(present_entry.get() or 0)
                if total > 0 or present > 0:
                    total_number_of_staff.append(f"{label}: {total}")
                    no_of_presents.append(f"{label}: {present}")
                    staff_present += present
                    total_staff += total
            except ValueError:
                continue
        
        total_Staff_Absent = total_staff - staff_present
    
        # Format student attendance
        total_student_attendance = []
        total_student_presents = []
        
        try:
            preprimary_present = int(self.preprimary_present.get() or 0)
            total_preprimary = int(self.total_preprimary.get() or 0)
            if total_preprimary > 0 or preprimary_present > 0:
                total_student_attendance.append(f"PP- {total_preprimary}")
                total_student_presents.append(f"PP- {preprimary_present}")
        except ValueError:
            pass
        
        try:
            primary_present = int(self.primary_present.get() or 0)
            total_primary = int(self.total_primary.get() or 0)
            if total_primary > 0 or primary_present > 0:
                total_student_attendance.append(f"PR- {total_primary}")
                total_student_presents.append(f"PR- {primary_present}")
        except ValueError:
            pass
        
        total_present = sum(map(int, [self.preprimary_present.get() or 0, self.primary_present.get() or 0]))
        total_students = sum(map(int, [self.total_preprimary.get() or 0, self.total_primary.get() or 0]))
        total_absent = total_students - total_present
        
        total_student_presents.append(f"Total- {total_present}")
    
        # Add data to row
        data = [
            "\n".join(day_date),
            "\n".join(total_number_of_staff),
            "\n".join(no_of_presents),
            total_Staff_Absent,
            "\n".join(total_student_attendance),
            "\n".join(total_student_presents),
            total_absent,
            self.remarks_text.get("1.0", tk.END).strip(),
            self.issues_text.get("1.0", tk.END).strip()
        ]
    
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=next_row, column=col)
            cell.value = value
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            cell.border = thick_border  # Apply thick border
    
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap width at 50 characters
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
        # Save the workbook
        try:
            wb.save(full_path)
            messagebox.showinfo("Success", f"Data exported to {full_path}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export data: {str(e)}")
            

if __name__ == "__main__":
    root = tk.Tk()
    app = MISCreator(root)
    root.mainloop()